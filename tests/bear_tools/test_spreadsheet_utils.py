"""
Comprehensive unit tests for bear_tools.spreadsheet_utils module.

This module contains type-hinted unit tests for the SpreadsheetBuilder class
and related utilities, including chart creation, data formatting, and file operations.
"""
# pylint: disable=protected-access  # Accessing protected members is acceptable in tests

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
from openpyxl.workbook import Workbook

from bear_tools.spreadsheet_utils import ChartConfig, ChartType, SpreadsheetBuilder, create_basic_spreadsheet


class TestChartType:
    """Test cases for ChartType enum."""

    def test_chart_type_values(self) -> None:
        """Test that ChartType enum has correct values."""
        assert ChartType.PIE.value == "pie"
        assert ChartType.SCATTER.value == "scatter"
        assert ChartType.LINE.value == "line"
        assert ChartType.BAR.value == "bar"
        assert ChartType.COLUMN.value == "column"

    def test_chart_type_members(self) -> None:
        """Test that all expected ChartType members exist."""
        expected_members = {"PIE", "SCATTER", "LINE", "BAR", "COLUMN"}
        actual_members = {member.name for member in ChartType}
        assert actual_members == expected_members


class TestChartConfig:
    """Test cases for ChartConfig dataclass."""

    def test_default_values(self) -> None:
        """Test ChartConfig default values."""
        config = ChartConfig()
        assert config.title is None
        assert config.x_axis_title is None
        assert config.y_axis_title is None
        assert config.show_legend is True
        assert config.show_data_labels is False
        assert config.width == 15
        assert config.height == 10

    def test_custom_values(self) -> None:
        """Test ChartConfig with custom values."""
        config = ChartConfig(
            title="Test Chart",
            x_axis_title="X Axis",
            y_axis_title="Y Axis",
            show_legend=False,
            show_data_labels=True,
            width=20,
            height=15
        )
        assert config.title == "Test Chart"
        assert config.x_axis_title == "X Axis"
        assert config.y_axis_title == "Y Axis"
        assert config.show_legend is False
        assert config.show_data_labels is True
        assert config.width == 20
        assert config.height == 15

    def test_partial_custom_values(self) -> None:
        """Test ChartConfig with some custom values."""
        config = ChartConfig(title="My Chart", show_legend=False)
        assert config.title == "My Chart"
        assert config.show_legend is False
        # Check defaults are still applied
        assert config.width == 15
        assert config.height == 10


class TestSpreadsheetBuilder:
    """Test cases for SpreadsheetBuilder class."""

    @pytest.fixture
    def temp_output_path(self, tmp_path: Path) -> Path:
        """Create a temporary output path for testing."""
        return tmp_path / "test_spreadsheet.xlsx"

    @pytest.fixture
    def builder(self, temp_output_path: Path) -> SpreadsheetBuilder:
        """Create a SpreadsheetBuilder instance for testing."""
        return SpreadsheetBuilder(temp_output_path)

    @pytest.fixture
    def sample_data(self) -> list[list[int | float | str]]:
        """Sample data for testing."""
        return [
            ["Product", "Q1", "Q2"],
            ["Widget", 100, 150],
            ["Gadget", 200, 180],
        ]

    def test_init(self, temp_output_path: Path) -> None:
        """Test SpreadsheetBuilder initialization."""
        builder = SpreadsheetBuilder(temp_output_path)
        assert builder._output == temp_output_path
        assert isinstance(builder._workbook, Workbook)
        assert len(builder._sheets) == 0
        # Default sheet should be removed
        assert len(builder._workbook.worksheets) == 0

    def test_add_sheet_basic(self, builder: SpreadsheetBuilder, sample_data: list[list[int | float | str]]) -> None:
        """Test adding a basic sheet with data."""
        result = builder.add_sheet("Sales", sample_data)

        # Check method chaining
        assert result is builder

        # Check sheet was created
        assert "Sales" in builder._sheets
        ws = builder._sheets["Sales"]

        # Check data was written
        assert ws["A1"].value == "Product"
        assert ws["B1"].value == "Q1"
        assert ws["C1"].value == "Q2"
        assert ws["A2"].value == "Widget"
        assert ws["B2"].value == 100
        assert ws["C2"].value == 150

    def test_add_sheet_header_formatting(
        self, builder: SpreadsheetBuilder, sample_data: list[list[int | float | str]]
    ) -> None:
        """Test that header row is formatted correctly."""
        builder.add_sheet("Sales", sample_data, first_row_is_header=True)
        ws = builder._sheets["Sales"]

        # Check header formatting
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 14
        assert ws["B1"].font.bold is True
        assert ws["C1"].font.bold is True

    def test_add_sheet_no_header_formatting(
        self, builder: SpreadsheetBuilder, sample_data: list[list[int | float | str]]
    ) -> None:
        """Test that header formatting is skipped when first_row_is_header=False."""
        builder.add_sheet("Sales", sample_data, first_row_is_header=False)
        ws = builder._sheets["Sales"]

        # Check no special formatting (no bold, default size)
        assert ws["A1"].font.bold is not True
        assert ws["A1"].font.size != 14  # Not the header font size

    def test_add_sheet_empty_data(self, builder: SpreadsheetBuilder) -> None:
        """Test adding a sheet with empty data."""
        empty_data: list[list[int | float | str]] = []
        builder.add_sheet("Empty", empty_data)

        assert "Empty" in builder._sheets
        ws = builder._sheets["Empty"]
        # Sheet should exist but be empty
        assert ws.max_row == 1  # openpyxl always has at least 1 row

    def test_add_sheet_mixed_types(self, builder: SpreadsheetBuilder) -> None:
        """Test adding a sheet with mixed data types."""
        mixed_data: list[list[int | float | str]] = [
            ["Name", "Age", "Score", "Status"],
            ["Alice", 25, 95.5, "Pass"],
            ["Bob", 30, 87.3, "Pass"],
        ]
        builder.add_sheet("Mixed", mixed_data)
        ws = builder._sheets["Mixed"]

        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 25
        assert ws["C2"].value == 95.5
        assert ws["D2"].value == "Pass"

    def test_add_multiple_sheets(self, builder: SpreadsheetBuilder, sample_data: list[list[int | float | str]]) -> None:
        """Test adding multiple sheets."""
        builder.add_sheet("Sales", sample_data)
        builder.add_sheet("Customers", [["Name", "Count"], ["Alice", 10]])

        assert len(builder._sheets) == 2
        assert "Sales" in builder._sheets
        assert "Customers" in builder._sheets

    def test_add_pie_chart(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a pie chart."""
        data: list[list[int | float | str]] = [
            ["Product", "Sales"],
            ["Widget", 100],
            ["Gadget", 200],
        ]
        builder.add_sheet("Data", data)
        result = builder.add_pie_chart(
            sheet_name="Data",
            data_range="A1:B3",
            anchor_cell="D2",
            config=ChartConfig(title="Sales by Product")
        )

        # Check method chaining
        assert result is builder

        # Save and verify chart exists
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1  # type: ignore[attr-defined]
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert isinstance(chart, PieChart)
        assert chart.title.text.rich.p[0].r[0].t == "Sales by Product"  # type: ignore[union-attr]

    def test_add_pie_chart_invalid_sheet(self, builder: SpreadsheetBuilder) -> None:
        """Test adding pie chart to non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet 'NonExistent' does not exist"):
            builder.add_pie_chart("NonExistent", "A1:B3", "D2")

    def test_add_scatter_chart(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a scatter chart."""
        data: list[list[int | float | str]] = [
            ["X", "Y"],
            [10, 100],
            [20, 200],
            [30, 300],
        ]
        builder.add_sheet("Data", data)
        result = builder.add_scatter_chart(
            sheet_name="Data",
            x_range="A2:A4",
            y_range="B2:B4",
            anchor_cell="D2",
            config=ChartConfig(title="Scatter Plot", x_axis_title="X Values", y_axis_title="Y Values"),
            series_name="Test Series"
        )

        # Check method chaining
        assert result is builder

        # Save and verify chart exists
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1  # type: ignore[attr-defined]
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert isinstance(chart, ScatterChart)
        assert chart.title.text.rich.p[0].r[0].t == "Scatter Plot"  # type: ignore[union-attr]
        assert chart.x_axis.title.text.rich.p[0].r[0].t == "X Values"  # type: ignore[union-attr]
        assert chart.y_axis.title.text.rich.p[0].r[0].t == "Y Values"  # type: ignore[union-attr]

        # Verify marker configuration
        assert len(chart.series) == 1
        series = chart.series[0]
        assert series.marker is not None
        assert series.marker.symbol == 'circle'
        assert series.marker.size == 10
        assert series.smooth is False

    def test_add_scatter_chart_invalid_sheet(self, builder: SpreadsheetBuilder) -> None:
        """Test adding scatter chart to non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet 'NonExistent' does not exist"):
            builder.add_scatter_chart("NonExistent", "A2:A4", "B2:B4", "D2")

    def test_add_line_chart(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a line chart."""
        data: list[list[int | float | str]] = [
            ["Month", "Sales", "Expenses"],
            ["Jan", 100, 80],
            ["Feb", 120, 90],
            ["Mar", 140, 95],
        ]
        builder.add_sheet("Data", data)
        result = builder.add_line_chart(
            sheet_name="Data",
            data_range="B1:C4",
            categories_range="A2:A4",
            anchor_cell="E2",
            config=ChartConfig(title="Monthly Trends")
        )

        # Check method chaining
        assert result is builder

        # Save and verify chart exists
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1  # type: ignore[attr-defined]
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert isinstance(chart, LineChart)
        assert chart.title.text.rich.p[0].r[0].t == "Monthly Trends"  # type: ignore[union-attr]

    def test_add_line_chart_invalid_sheet(self, builder: SpreadsheetBuilder) -> None:
        """Test adding line chart to non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet 'NonExistent' does not exist"):
            builder.add_line_chart("NonExistent", "B1:C4", "A2:A4", "E2")

    def test_add_bar_chart_horizontal(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a horizontal bar chart."""
        data: list[list[int | float | str]] = [
            ["Product", "Q1"],
            ["Widget", 100],
            ["Gadget", 200],
        ]
        builder.add_sheet("Data", data)
        result = builder.add_bar_chart(
            sheet_name="Data",
            data_range="B1:B3",
            categories_range="A2:A3",
            anchor_cell="D2",
            config=ChartConfig(title="Horizontal Bar Chart"),
            horizontal=True
        )

        # Check method chaining
        assert result is builder

        # Save and verify chart exists
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1  # type: ignore[attr-defined]
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert isinstance(chart, BarChart)
        assert chart.type == "bar"

    def test_add_bar_chart_vertical(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a vertical bar chart (column chart)."""
        data: list[list[int | float | str]] = [
            ["Product", "Q1"],
            ["Widget", 100],
            ["Gadget", 200],
        ]
        builder.add_sheet("Data", data)
        builder.add_bar_chart(
            sheet_name="Data",
            data_range="B1:B3",
            categories_range="A2:A3",
            anchor_cell="D2",
            config=ChartConfig(title="Column Chart"),
            horizontal=False
        )

        # Save and verify chart exists
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1  # type: ignore[attr-defined]
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert isinstance(chart, BarChart)
        assert chart.type == "col"

    def test_add_bar_chart_invalid_sheet(self, builder: SpreadsheetBuilder) -> None:
        """Test adding bar chart to non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet 'NonExistent' does not exist"):
            builder.add_bar_chart("NonExistent", "B1:B3", "A2:A3", "D2")

    def test_add_hyperlink(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding a hyperlink to a cell."""
        data: list[list[int | float | str]] = [["Name", "URL"], ["Example", ""]]
        builder.add_sheet("Links", data)
        result = builder.add_hyperlink(
            sheet_name="Links",
            cell="B2",
            url="https://example.com",
            display_text="Visit Example"
        )

        # Check method chaining
        assert result is builder

        # Save and verify hyperlink
        builder.save()
        wb = load_workbook(temp_output_path)
        ws = wb["Links"]
        cell = ws["B2"]
        assert cell.value == "Visit Example"
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == "https://example.com"
        # Check hyperlink styling (blue color and underline)
        assert cell.font.color is not None  # Has color
        assert cell.font.underline == "single"

    def test_add_hyperlink_invalid_sheet(self, builder: SpreadsheetBuilder) -> None:
        """Test adding hyperlink to non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet 'NonExistent' does not exist"):
            builder.add_hyperlink("NonExistent", "B2", "https://example.com")

    def test_save_success(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test successfully saving a spreadsheet."""
        builder.add_sheet("Test", [["A", "B"], [1, 2]])
        result = builder.save()

        assert result is True
        assert temp_output_path.exists()

        # Verify file can be loaded
        wb = load_workbook(temp_output_path)
        assert "Test" in wb.sheetnames

    def test_save_creates_parent_directory(self, tmp_path: Path) -> None:
        """Test that save() creates parent directories if they don't exist."""
        nested_path = tmp_path / "subdir" / "nested" / "test.xlsx"
        builder = SpreadsheetBuilder(nested_path)
        builder.add_sheet("Test", [["Data"]])

        result = builder.save()

        assert result is True
        assert nested_path.exists()
        assert nested_path.parent.exists()

    def test_save_empty_workbook(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test saving a workbook with at least one empty sheet."""
        # openpyxl requires at least one sheet, so add a minimal empty sheet
        builder.add_sheet("Empty", [[]])
        result = builder.save()
        assert result is True
        assert temp_output_path.exists()

    def test_method_chaining(self, builder: SpreadsheetBuilder, sample_data: list[list[int | float | str]]) -> None:
        """Test that all builder methods support chaining."""
        result = (
            builder
            .add_sheet("Sales", sample_data)
            .add_pie_chart("Sales", "A1:B3", "D2")
            .add_hyperlink("Sales", "E1", "https://example.com", "Link")
        )

        assert result is builder
        assert len(builder._sheets) == 1
        assert "Sales" in builder._sheets

    def test_chart_config_applied_to_pie_chart(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test that ChartConfig is properly applied to pie charts."""
        data: list[list[int | float | str]] = [["Item", "Value"], ["A", 10], ["B", 20]]
        builder.add_sheet("Data", data)
        builder.add_pie_chart(
            "Data",
            "A1:B3",
            "D2",
            ChartConfig(title="Test Title", show_legend=False, show_data_labels=True)
        )
        builder.save()

        wb = load_workbook(temp_output_path)
        chart = wb["Data"]._charts[0]  # type: ignore[attr-defined]
        assert chart.title.text.rich.p[0].r[0].t == "Test Title"  # type: ignore[union-attr]
        # Note: Testing legend/data labels requires deeper inspection of chart properties

    def test_multiple_charts_on_same_sheet(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test adding multiple charts to the same sheet."""
        data: list[list[int | float | str]] = [
            ["X", "Y", "Z"],
            [1, 10, 100],
            [2, 20, 200],
            [3, 30, 300],
        ]
        builder.add_sheet("Data", data)
        builder.add_scatter_chart("Data", "A2:A4", "B2:B4", "E2", ChartConfig(title="Chart 1"))
        builder.add_scatter_chart("Data", "A2:A4", "C2:C4", "E15", ChartConfig(title="Chart 2"))
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Data"]
        assert len(ws._charts) == 2  # type: ignore[attr-defined]

    def test_unicode_in_data(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test handling of Unicode characters in data."""
        unicode_data: list[list[int | float | str]] = [
            ["Name", "Country"],
            ["François", "France 🇫🇷"],
            ["日本", "Japan"],
            ["Москва", "Russia"],
        ]
        builder.add_sheet("Unicode", unicode_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Unicode"]
        assert ws["A2"].value == "François"
        assert ws["B2"].value == "France 🇫🇷"
        assert ws["A3"].value == "日本"

    def test_large_dataset(self, builder: SpreadsheetBuilder, temp_output_path: Path) -> None:
        """Test handling of a large dataset."""
        # Create dataset with 1000 rows
        large_data: list[list[int | float | str]] = [["Index", "Value"]]
        for i in range(1000):
            large_data.append([i, i * 2.5])

        builder.add_sheet("Large", large_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Large"]
        assert ws.max_row == 1001  # Header + 1000 rows
        assert ws["A1001"].value == 999
        assert ws["B1001"].value == 999 * 2.5


class TestCreateBasicSpreadsheet:
    """Test cases for the create_basic_spreadsheet convenience function."""

    @pytest.fixture
    def temp_output_path(self, tmp_path: Path) -> Path:
        """Create a temporary output path for testing."""
        return tmp_path / "basic_test.xlsx"

    def test_create_basic_spreadsheet(self, temp_output_path: Path) -> None:
        """Test creating a basic spreadsheet with the convenience function."""
        data: list[list[int | float | str]] = [
            ["Name", "Age"],
            ["Alice", 25],
            ["Bob", 30],
        ]

        result = create_basic_spreadsheet(data, temp_output_path, title="People")

        assert result is True
        assert temp_output_path.exists()

        # Verify contents
        wb = load_workbook(temp_output_path)
        assert "People" in wb.sheetnames
        ws = wb["People"]
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Age"
        assert ws["A2"].value == "Alice"

    def test_create_basic_spreadsheet_default_title(self, temp_output_path: Path) -> None:
        """Test creating a basic spreadsheet with default title."""
        data: list[list[int | float | str]] = [["A", "B"], [1, 2]]

        result = create_basic_spreadsheet(data, temp_output_path)

        assert result is True
        wb = load_workbook(temp_output_path)
        assert "Data" in wb.sheetnames


class TestEdgeCases:
    """Test edge cases and error handling."""

    @pytest.fixture
    def temp_output_path(self, tmp_path: Path) -> Path:
        """Create a temporary output path for testing."""
        return tmp_path / "edge_case_test.xlsx"

    def test_single_row_data(self, temp_output_path: Path) -> None:
        """Test handling of single-row data."""
        builder = SpreadsheetBuilder(temp_output_path)
        single_row: list[list[int | float | str]] = [["Only", "Header", "Row"]]
        builder.add_sheet("Single", single_row)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Single"]
        assert ws.max_row == 1
        assert ws["A1"].value == "Only"

    def test_single_column_data(self, temp_output_path: Path) -> None:
        """Test handling of single-column data."""
        builder = SpreadsheetBuilder(temp_output_path)
        single_col: list[list[int | float | str]] = [["A"], [1], [2], [3]]
        builder.add_sheet("Column", single_col)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Column"]
        assert ws.max_column == 1
        assert ws["A4"].value == 3

    def test_very_long_sheet_name(self, temp_output_path: Path) -> None:
        """Test handling of very long sheet names."""
        builder = SpreadsheetBuilder(temp_output_path)
        # Excel has a 31-character limit on sheet names
        long_name = "A" * 31
        builder.add_sheet(long_name, [["Data"]])
        builder.save()

        wb = load_workbook(temp_output_path)
        # openpyxl should handle this, possibly truncating
        assert len(wb.sheetnames) > 0

    def test_special_characters_in_sheet_name(self, temp_output_path: Path) -> None:
        """Test handling of special characters in sheet names."""
        builder = SpreadsheetBuilder(temp_output_path)
        # Some characters are invalid in Excel sheet names: [ ] : * ? / \
        # Test with valid special characters
        special_name = "Data-2024_v1.0 (Final)"
        builder.add_sheet(special_name, [["Test"]])
        builder.save()

        wb = load_workbook(temp_output_path)
        assert special_name in wb.sheetnames

    def test_numeric_only_data(self, temp_output_path: Path) -> None:
        """Test handling of data with only numeric values."""
        builder = SpreadsheetBuilder(temp_output_path)
        numeric_data: list[list[int | float | str]] = [
            [1, 2, 3],
            [4, 5, 6],
            [7.5, 8.9, 10.2],
        ]
        builder.add_sheet("Numbers", numeric_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Numbers"]
        assert ws["A3"].value == 7.5
        assert ws["C3"].value == 10.2

    def test_zero_values(self, temp_output_path: Path) -> None:
        """Test handling of zero values in data."""
        builder = SpreadsheetBuilder(temp_output_path)
        zero_data: list[list[int | float | str]] = [
            ["Value"],
            [0],
            [0.0],
            [0],
        ]
        builder.add_sheet("Zeros", zero_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Zeros"]
        assert ws["A2"].value == 0
        assert ws["A3"].value == 0.0

    def test_negative_values(self, temp_output_path: Path) -> None:
        """Test handling of negative values."""
        builder = SpreadsheetBuilder(temp_output_path)
        negative_data: list[list[int | float | str]] = [
            ["Value"],
            [-100],
            [-3.14],
        ]
        builder.add_sheet("Negative", negative_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Negative"]
        assert ws["A2"].value == -100
        assert ws["A3"].value == -3.14

    def test_empty_string_values(self, temp_output_path: Path) -> None:
        """Test handling of empty string values."""
        builder = SpreadsheetBuilder(temp_output_path)
        empty_data: list[list[int | float | str]] = [
            ["Name", "Value"],
            ["", 10],
            ["Item", ""],
        ]
        builder.add_sheet("Empty", empty_data)
        builder.save()

        wb = load_workbook(temp_output_path)
        ws = wb["Empty"]
        # openpyxl treats empty strings as None when saving/loading
        assert ws["A2"].value is None or ws["A2"].value == ""
        assert ws["B3"].value is None or ws["B3"].value == ""


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
