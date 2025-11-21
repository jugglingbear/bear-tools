"""
Module designed to streamline certain common/repetitive spreadsheet-related activities
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.series_factory import SeriesFactory as Series
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from bear_tools import lumberjack

logger = lumberjack.Logger()



@dataclass
class ChartConfig:
    """Configuration for chart appearance and behavior"""
    title: str | None = None
    x_axis_title: str | None = None
    y_axis_title: str | None = None
    show_legend: bool = True
    show_data_labels: bool = False
    width: int = 15  # Width in Excel columns
    height: int = 10  # Height in Excel rows


class SpreadsheetBuilder:
    """
    Builder class for creating Excel spreadsheets with data, formatting, charts, and hyperlinks.

    Example:
        >>> builder = SpreadsheetBuilder(output=Path("report.xlsx"))
        >>> builder.add_sheet(
        ...     name="Sales Data",
        ...     data=[["Product", "Q1", "Q2"], ["Widget", 100, 150], ["Gadget", 200, 180]],
        ...     first_row_is_header=True
        ... )
        >>> builder.add_pie_chart(
        ...     sheet_name="Sales Data",
        ...     data_range="A1:B3",
        ...     anchor_cell="D2",
        ...     config=ChartConfig(title="Q1 Sales by Product")
        ... )
        >>> builder.save()
    """

    def __init__(self, output: Path) -> None:
        """
        Initialize a new spreadsheet builder

        :param output: The path where the spreadsheet will be saved
        """
        self._output: Path = output
        self._workbook: Workbook = Workbook()
        # Remove default sheet
        if self._workbook.active is not None:
            self._workbook.remove(self._workbook.active)
        self._sheets: dict[str, Worksheet] = {}

    def add_sheet(
        self,
        name: str,
        data: list[list[int | float | str]],
        first_row_is_header: bool = True,
        autofit_columns: bool = True
    ) -> SpreadsheetBuilder:
        """
        Add a worksheet with data to the spreadsheet

        :param name: The name of the worksheet (tab name)
        :param data: 2D list of data to populate the sheet
        :param first_row_is_header: If True, format the first row as a header (bold, larger font)
        :param autofit_columns: If True, automatically adjust column widths to fit content
        :return: Self for method chaining
        """
        ws: Worksheet = self._workbook.create_sheet(title=name)
        self._sheets[name] = ws

        # Write data
        df = pandas.DataFrame(data)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Format header row
        if first_row_is_header and len(data) > 0:
            for i in range(len(data[0])):
                col_letter = get_column_letter(i + 1)
                ws[f'{col_letter}1'].font = Font(bold=True, size=14)

        # Autofit columns
        if autofit_columns:
            self._autofit_columns(ws)

        return self

    def add_pie_chart(
        self,
        sheet_name: str,
        data_range: str,
        anchor_cell: str,
        config: ChartConfig | None = None
    ) -> SpreadsheetBuilder:
        """
        Add a pie chart to a worksheet

        :param sheet_name: Name of the worksheet to add the chart to
        :param data_range: Excel range containing the data in format "A1:B5" where column A has labels, B has values
        :param anchor_cell: Cell where the top-left corner of the chart will be placed (e.g., "D2")
        :param config: Optional chart configuration
        :return: Self for method chaining
        """

        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        ws = self._sheets[sheet_name]
        config = config or ChartConfig()

        chart = PieChart()
        self._apply_chart_config(chart, config)

        # Parse the data range to extract label and value columns
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f"Invalid cell range: {data_range}")

        # For pie charts: first column is labels, second column is values
        # Values reference (column B in typical "A1:B5" range)
        values_col = get_column_letter(max_col)
        data_ref = self._create_reference(sheet_name, f"{values_col}{min_row}:{values_col}{max_row}")
        chart.add_data(data_ref, titles_from_data=True)

        # Labels reference (column A, excluding header)
        if max_row > min_row:  # Only if there's more than just header
            labels_col = get_column_letter(min_col)
            labels_ref = self._create_reference(sheet_name, f"{labels_col}{min_row + 1}:{labels_col}{max_row}")
            chart.set_categories(labels_ref)

        # Set chart size
        chart.width = config.width
        chart.height = config.height

        ws.add_chart(chart, anchor_cell)
        return self

    def add_scatter_chart(
        self,
        sheet_name: str,
        x_range: str,
        y_range: str,
        anchor_cell: str,
        config: ChartConfig | None = None,
        series_name: str | None = None
    ) -> SpreadsheetBuilder:
        """
        Add a scatter plot to a worksheet

        :param sheet_name: Name of the worksheet to add the chart to
        :param x_range: Excel range for X-axis data (e.g., "A2:A10")
        :param y_range: Excel range for Y-axis data (e.g., "B2:B10")
        :param anchor_cell: Cell where the top-left corner of the chart will be placed (e.g., "D2")
        :param config: Optional chart configuration
        :param series_name: Optional name for the data series
        :return: Self for method chaining
        """
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        ws = self._sheets[sheet_name]
        config = config or ChartConfig()

        chart = ScatterChart()
        # Do NOT set scatterStyle - it prevents markers from showing in Excel
        self._apply_chart_config(chart, config)

        # Create data references
        x_values = self._create_reference(sheet_name, x_range)
        y_values = self._create_reference(sheet_name, y_range)

        # Create series manually with marker configuration for Excel compatibility

        series = Series(values=y_values, xvalues=x_values)
        marker = Marker(symbol='circle')
        marker.size = 10  # Visible marker size
        series.marker = marker
        series.smooth = False  # Disable smoothing for clearer markers

        chart.series.append(series)

        # Set series name if provided
        if series_name:
            chart.series[0].title = SeriesLabel(v=series_name)

        # Set chart size
        chart.width = config.width
        chart.height = config.height

        ws.add_chart(chart, anchor_cell)
        return self

    def add_line_chart(
        self,
        sheet_name: str,
        data_range: str,
        categories_range: str | None,
        anchor_cell: str,
        config: ChartConfig | None = None
    ) -> SpreadsheetBuilder:
        """
        Add a line chart to a worksheet

        :param sheet_name: Name of the worksheet to add the chart to
        :param data_range: Excel range for Y-axis data series (e.g., "B1:C10")
        :param categories_range: Optional Excel range for X-axis categories (e.g., "A2:A10")
        :param anchor_cell: Cell where the top-left corner of the chart will be placed (e.g., "E2")
        :param config: Optional chart configuration
        :return: Self for method chaining
        """
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        ws = self._sheets[sheet_name]
        config = config or ChartConfig()

        chart = LineChart()
        self._apply_chart_config(chart, config)

        # Add data
        data_ref = self._create_reference(sheet_name, data_range)
        chart.add_data(data_ref, titles_from_data=True)

        # Add categories if provided
        if categories_range:
            cats_ref = self._create_reference(sheet_name, categories_range)
            chart.set_categories(cats_ref)

        # Set chart size
        chart.width = config.width
        chart.height = config.height

        ws.add_chart(chart, anchor_cell)
        return self

    def add_bar_chart(
        self,
        sheet_name: str,
        data_range: str,
        categories_range: str | None,
        anchor_cell: str,
        config: ChartConfig | None = None,
        horizontal: bool = True
    ) -> SpreadsheetBuilder:
        """
        Add a bar or column chart to a worksheet

        :param sheet_name: Name of the worksheet to add the chart to
        :param data_range: Excel range for data series (e.g., "B1:C10")
        :param categories_range: Optional Excel range for categories (e.g., "A2:A10")
        :param anchor_cell: Cell where the top-left corner of the chart will be placed (e.g., "E2")
        :param config: Optional chart configuration
        :param horizontal: If True, create horizontal bars; if False, create vertical columns
        :return: Self for method chaining
        """
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        ws = self._sheets[sheet_name]
        config = config or ChartConfig()

        chart = BarChart()
        chart.type = "bar" if horizontal else "col"
        self._apply_chart_config(chart, config)

        # Add data
        data_ref = self._create_reference(sheet_name, data_range)
        chart.add_data(data_ref, titles_from_data=True)

        # Add categories if provided
        if categories_range:
            cats_ref = self._create_reference(sheet_name, categories_range)
            chart.set_categories(cats_ref)

        # Set chart size
        chart.width = config.width
        chart.height = config.height

        ws.add_chart(chart, anchor_cell)
        return self

    def add_hyperlink(
        self,
        sheet_name: str,
        cell: str,
        url: str,
        display_text: str | None = None
    ) -> SpreadsheetBuilder:
        """
        Add a hyperlink to a cell

        :param sheet_name: Name of the worksheet
        :param cell: Cell reference (e.g., "A1")
        :param url: URL to link to (e.g., "http://www.google.com")
        :param display_text: Optional text to display (defaults to URL)
        :return: Self for method chaining
        """
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        ws = self._sheets[sheet_name]
        target_cell = ws[cell]

        # Set the hyperlink
        target_cell.hyperlink = url

        # Set display text
        if display_text:
            target_cell.value = display_text
        elif not target_cell.value:
            target_cell.value = url

        # Style as a link (blue, underlined)
        target_cell.font = Font(color="0563C1", underline="single")

        return self

    def save(self) -> bool:
        """
        Save the workbook to the output file

        :return: True if save was successful, False otherwise
        """
        try:
            self._output.parent.mkdir(parents=True, exist_ok=True)
            self._workbook.save(self._output)
            logger.info(f'Generated spreadsheet: {self._output}', color=lumberjack.PrintColor.GREEN)
            return True
        except (OSError, PermissionError) as e:
            logger.error(f'Failed to save spreadsheet: {e}')
            return False

    def _apply_chart_config(self, chart: PieChart | ScatterChart | LineChart | BarChart, config: ChartConfig) -> None:
        """
        Apply configuration to a chart

        :param chart: The chart object to configure
        :param config: Configuration to apply
        """
        if config.title:
            chart.title = config.title

        # Apply axis titles (not applicable for pie charts)
        if not isinstance(chart, PieChart):
            if hasattr(chart, 'x_axis') and config.x_axis_title:
                chart.x_axis.title = config.x_axis_title
            if hasattr(chart, 'y_axis') and config.y_axis_title:
                chart.y_axis.title = config.y_axis_title

        # Configure legend
        if not config.show_legend:
            chart.legend = None

        # Configure data labels
        if config.show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

    def _autofit_columns(self, ws: Worksheet) -> None:
        """
        Automatically adjust column widths to fit content

        :param ws: Worksheet to adjust
        """
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            column_idx = col[0].column
            if column_idx is not None:
                column_letter = get_column_letter(column_idx)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _create_reference(self, sheet_name: str, cell_range: str) -> Reference:
        """
        Create a Reference object with proper sheet name formatting

        :param sheet_name: Name of the worksheet
        :param cell_range: Cell range (e.g., "A1:B5")
        :return: Reference object
        """

        ws = self._sheets[sheet_name]
        # Parse the range (e.g., "A1:B5") into coordinates
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        # Ensure all values are not None
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f"Invalid cell range: {cell_range}")

        return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def create_basic_spreadsheet(
    data: list[list[int | float | str]],
    output: Path,
    title: str = 'Data',
    first_row_is_header: bool = True
) -> bool:
    """
    Create a basic spreadsheet from two-dimensional data (simple convenience function)

    :param data: A 2D table of data to write into a spreadsheet
    :param output: The path to the output file
    :param title: The title of the worksheet (i.e. the tab)
    :param first_row_is_header: If True, the first row in data is header (bold); otherwise, not
    :return: True if everything was successful, False otherwise
    """
    builder = SpreadsheetBuilder(output)
    builder.add_sheet(name=title, data=data, first_row_is_header=first_row_is_header)
    return builder.save()
